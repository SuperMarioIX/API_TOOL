from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def saveDataIntoRow(sheet, row_num, data):
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=row_num, column=col_num, value=value)

    for col_num in range(1, len(data) + 1):
        column_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in sheet[column_letter])
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width