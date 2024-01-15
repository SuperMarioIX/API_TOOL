import json
from openpyxl import Workbook
import SaveDataInExcel
from openpyxl.styles import PatternFill

def extract_info(result):
    try:
        pronto_id = result["pronto_id"]
        repeatability = result["repeatability"]
        tribe_name = result["group_in_charge_tribe_name"]
        reported_date = result["reported_date"]
        closed_date = result.get("closed_date", None)  # Folosim get pentru a gestiona absența cheii
        title = result["title"]
        group_in_charge_name = result["group_in_charge_name"]
        state = result["state"]
        feature = result["feature"]

        return [feature, pronto_id, state, repeatability, reported_date, closed_date, tribe_name,group_in_charge_name, title]

    except KeyError as e:
        print(f"KeyError: {e} is missing in the data.")
        return []

# Citirea datelor din fișierul output.json
def openJsonFile():
    with open('output.json', 'r') as file:
        data = json.load(file)
    return data

def fillFirstCoLFromExcel(sheet):
    headers = ["FEATURE", "PRONTO_ID", "STATUS", "REPEATABILITY", "REPORTED DATE", "CLOSED DATE", "TRIBE OF GROUP IN CHARGE", "GROUP IN CHARGE", "TITLE"]

# Fill the first column with headers
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)

# Apply violet color to the first row
    violet_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = violet_fill

def extractDataFromJsonAndWriteIntoExcel(feature_ID, pr_numbers):
    excel_filename =str(feature_ID) + '_' + 'prCount' + str(pr_numbers) + '.xlsx'

    data = openJsonFile()
    if 'results' in data:
        results = data['results']
        workbook = Workbook()
        sheet = workbook.active
        fillFirstCoLFromExcel(sheet)
        
        for col_num, result in enumerate(results, start=1):
            if result:
                # Extrage informațiile din primul element
                info_list = extract_info(result)
                # print(info_list)
                SaveDataInExcel.saveDataIntoRow(sheet, col_num+1, info_list)
            else:
                print("No results found in the data.")
        workbook.save(excel_filename)        
    else:
        print("Key 'results' not found in the data.")
